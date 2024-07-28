from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from time import sleep
from datetime import datetime
import openpyxl.workbook
import schedule 

# Iniciando o webdriver
def iniciar_driver():
    chrome_options = Options()
    arguments = ['--lang=pt-BR', '--window-size=800,600','--incognito']
    for argument in arguments:
        chrome_options.add_argument(argument)
    driver = webdriver.Chrome(options=chrome_options)
    return driver

link = "https://www.tumelero.com.br/resultado-busca?terms=cimento"

# Coletar os dados
def coletar_dados():
    driver = iniciar_driver()
    driver.get(link)
    sleep(2)
    preço = driver.find_element(By.XPATH, '//*[@id="vitrine-products-smarthint"]/div[1]/ul/li[1]/article/div[3]/div[3]/p/span[1]/strong')
    preço_baixo = preço.text
    sleep(2)
    driver.quit()
    return preço_baixo

# Criando a planilha onde vão ser guardados os dados
wb = openpyxl.Workbook()
ws = wb.active

# Criando cabeçalho da planilha
cabeçalho = ["Data", "Hora", "Produto", "Preço"]
for col, value in enumerate(cabeçalho, start = 1 ):
    ws.cell(row=1, column=col).value = value

# Configurando os dados
produto_nome = "Cimento"
preco = coletar_dados()
data_consulta = datetime.now().strftime('%Y-%m-%d')
hora_consulta = datetime.now().strftime('%H:%M:%S')

# Criar uma nova coluna com os novos dados criados
linha_dados = [data_consulta, hora_consulta, produto_nome, preco,]
for col, value in enumerate(linha_dados, start=1):
    ws.cell(row=2, column=col).value = value

# Salvar os dados coletados na planilha
wb.save("Preços do Cimento.xlsx")
    
# Agendamento para coleta de dados a cada 30 minutos
schedule.every(30).minutes.do(coletar_dados)
while True:
    schedule.run_pending()
    sleep(1)
